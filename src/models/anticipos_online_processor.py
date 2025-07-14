import pandas as pd
from typing import Dict, List
from src.models.data_models import AnticiposConfig
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class AnticiposOnlineProcessor:
    def __init__(self, config: AnticiposConfig = None):
        self.config = config if config else AnticiposConfig()
        
    def validate_input_file(self, file_path:str) -> bool:
        
        try:
            with pd.ExcelFile(file_path) as xls:
                sheets = xls.sheet_names
                missing_sheets = [sheet for sheet in self.config.required_sheets if sheet not in sheets]
                if missing_sheets:
                    raise ValueError(f"Faltan hojas requeridas: \n {', '.join(missing_sheets)}")
            return True
        except Exception as e:
            raise ValueError(f"Error al validar el archivo: \n {str(e)}")
        
    def load_and_filter_data(self, file_path: str) -> Dict[str, pd.DataFrame]:
        # ... (este método se queda igual) ...
        try:
            dfs = pd.read_excel(file_path,sheet_name=None)
            self.validate_input_file(file_path)
            filtered_data={}
            for sheet_name,columns in self.config.sheet_columns.items():
                if sheet_name not in dfs:
                    raise ValueError(f"Hoja {sheet_name} no encontrada en el archivo.\n")
                df = dfs[sheet_name][columns].copy()
                if sheet_name in self.config.rename_columns:
                    df.rename(columns= self.config.rename_columns[sheet_name], inplace=True)
                filtered_data[sheet_name] = df
            return filtered_data
        except Exception as e:
            raise ValueError(f"Error al cargar datos:\n {str(e)}")

    def save_formatted_excel(self, final_df: pd.DataFrame, output_filename: str):
        try:
            
            df_finansuenos = final_df[pd.notna(final_df['FACTURA_FS'])].copy()
            df_arpesod = final_df[pd.notna(final_df['FACTURA_ARP'])].copy()
            df_sin_cartera = final_df[(pd.isna(final_df['FACTURA_FS'])) & (pd.isna(final_df['FACTURA_ARP']))].copy()
            
            df_finansuenos_final = df_finansuenos.reindex(columns=self.config.column_order_fs)
            df_arpesod_final = df_arpesod.reindex(columns=self.config.column_order_arp)
            
            base_cols = list(self.config.rename_columns['ONLINE'].values())
            cols_sin_cartera = base_cols + ['CUENTAS_FS', 'CUENTAS_ARP', 'OBSERVACIONES']
            df_sin_cartera_final = df_sin_cartera.reindex(columns=cols_sin_cartera)
            
            
            dup_cedulas_fs = set(df_finansuenos_final[df_finansuenos_final.duplicated(subset=['CEDULA'], keep=False)]['CEDULA'])
            dup_facturas_fs = set(df_finansuenos_final[df_finansuenos_final.duplicated(subset=['FACTURA_FS'], keep=False)]['FACTURA_FS'])
            dup_cedulas_arp = set(df_arpesod_final[df_arpesod_final.duplicated(subset=['CEDULA'], keep=False)]['CEDULA'])
            dup_facturas_arp = set(df_arpesod_final[df_arpesod_final.duplicated(subset=['FACTURA_ARP'], keep=False)]['FACTURA_ARP'])

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                def format_sheet(df, sheet_name, duplicates_config=None):
                    df.fillna('', inplace=True)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]

                    
                    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    
                    
                    header_font = Font(bold=True, color="FFFFFF", name='Calibri')
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    header_alignment = Alignment(horizontal='center', vertical='center')
                    thin_border_side = Side(style='thin', color='000000')
                    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

                    col_indices = {cell.value: i for i, cell in enumerate(worksheet[1], 1)}
                    for cell in worksheet[1]: cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = cell_border
                    
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                        
                        # Obtener valores clave de la fila actual
                        factura_fs_val = row[col_indices.get('FACTURA_FS', 0) - 1] if 'FACTURA_FS' in col_indices else None
                        factura_arp_val = row[col_indices.get('FACTURA_ARP', 0) - 1] if 'FACTURA_ARP' in col_indices else None
                        obs_val = row[col_indices.get('OBSERVACIONES', 0) - 1] if 'OBSERVACIONES' in col_indices else None

                        # Regla 1 (Prioridad Alta): Si la factura indica múltiples carteras, pintar la fila de rojo.
                        if factura_fs_val == 'MAS DE UNA CARTERA' or factura_arp_val == 'MAS DE UNA CARTERA':
                            for cell in worksheet[row_idx]:
                                cell.fill = light_red_fill
                        
                        # Regla 2: Si la observación es 'PAGO TOTAL', pintar la fila de amarillo.
                        elif obs_val == 'PAGO TOTAL':
                            for cell in worksheet[row_idx]:
                                cell.fill = yellow_fill

                        # Regla 3: Duplicados de cédula o factura dentro de la misma hoja (pinta celdas específicas de rojo)
                        # Esto se mantiene para detectar otro tipo de duplicados.
                        if duplicates_config:
                            cedula_val = row[col_indices.get('CEDULA', 0) - 1] if 'CEDULA' in col_indices else None
                            if cedula_val in duplicates_config.get('cedulas', set()):
                                worksheet.cell(row=row_idx, column=col_indices['CEDULA']).fill = light_red_fill
                            
                            # Solo marcar si la factura no es el texto 'MAS DE UNA CARTERA'
                            if factura_fs_val and factura_fs_val != 'MAS DE UNA CARTERA' and factura_fs_val in duplicates_config.get('facturas', set()):
                                worksheet.cell(row=row_idx, column=col_indices['FACTURA_FS']).fill = light_red_fill
                            
                            if factura_arp_val and factura_arp_val != 'MAS DE UNA CARTERA' and factura_arp_val in duplicates_config.get('facturas', set()):
                                worksheet.cell(row=row_idx, column=col_indices['FACTURA_ARP']).fill = light_red_fill
                        
                        # Aplicar bordes a todas las celdas de la fila
                        for cell in worksheet[row_idx]:
                            cell.border = cell_border

                    # Autoajustar columnas (se mantiene igual)
                    for column in worksheet.columns:
                        max_length = 0; column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except: pass
                        adjusted_width = (max_length + 2); worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Las llamadas a la función se mantienen igual
                if not df_finansuenos_final.empty:
                    format_sheet(df_finansuenos_final, 'FINANSUEÑOS', duplicates_config={'cedulas': dup_cedulas_fs, 'facturas': dup_facturas_fs})
                if not df_arpesod_final.empty:
                    format_sheet(df_arpesod_final, 'ARPESOD', duplicates_config={'cedulas': dup_cedulas_arp, 'facturas': dup_facturas_arp})
                if not df_sin_cartera_final.empty:
                    format_sheet(df_sin_cartera_final, 'SIN CARTERA')

        except Exception as e:
            raise ValueError(f"Error al guardar el archivo Excel formateado:\n {str(e)}")