import pandas as pd
import numpy as np
from typing import Dict
from src.models.anticipos_model import AnticiposConfig
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class AnticiposService:
    def __init__(self, config: AnticiposConfig = None):
        self.config = config if config else AnticiposConfig()
        
    def generate_report(self, file_path: str, status_callback):
        """Orquesta el proceso completo de generación del reporte."""
        
        status_callback("Validando y cargando datos...", 10)
        dfs = self._load_and_filter_data(file_path)

        status_callback("Aplicando lógica de negocio...", 40)
        final_df = self._process_dataframes(dfs)

        status_callback("Preparando hojas finales...", 80)
        sheets_to_save = self._prepare_output_sheets(final_df)
        
        return sheets_to_save    
        
    def validate_input_file(self, file_path:str) -> bool:
        
        try:
            with pd.ExcelFile(file_path) as xls:
                sheets = xls.sheet_names
                missing_sheets = [sheet for sheet in self.config.required_sheets if sheet not in sheets]
                if missing_sheets:
                    raise ValueError(f"Faltan hojas requeridas: \n {', '.join(missing_sheets)}")
            return True
        except Exception as e:
            raise ValueError(f"Error al validar el archivo: \n {str(e)}")
        
    def _load_and_filter_data(self, file_path: str) -> Dict[str, pd.DataFrame]:
        # ... (este método se queda igual) ...
        try:
            dfs = pd.read_excel(file_path,sheet_name=None)
            self.validate_input_file(file_path)
            filtered_data={}
            for sheet_name,columns in self.config.sheet_columns.items():
                if sheet_name not in dfs:
                    raise ValueError(f"Hoja {sheet_name} no encontrada en el archivo.\n")
                df = dfs[sheet_name][columns].copy()
                if sheet_name in self.config.rename_columns:
                    df.rename(columns= self.config.rename_columns[sheet_name], inplace=True)
                filtered_data[sheet_name] = df
            return filtered_data
        except Exception as e:
            raise ValueError(f"Error al cargar datos:\n {str(e)}")
        
    def _process_dataframes(self, dfs: dict):
        # Lógica de tu antiguo `process_file` en el controlador
        df_online = dfs['ONLINE']
        df_ac_fs = dfs['AC FS']
        df_ac_arp = dfs['AC ARP']

        df_online['CEDULA'] = df_online['CEDULA'].astype(str).str.strip()
        df_ac_fs['CEDULA'] = df_ac_fs['CEDULA'].astype(str).str.strip()
        df_ac_arp['CEDULA'] = df_ac_arp['CEDULA'].astype(str).str.strip()

        df_online['CUENTAS_FS'] = df_online['CEDULA'].map(df_ac_fs['CEDULA'].value_counts()).fillna(0).astype(int)
        df_online['CUENTAS_ARP'] = df_online['CEDULA'].map(df_ac_arp['CEDULA'].value_counts()).fillna(0).astype(int)

        merged_df = pd.merge(df_online, df_ac_fs, on='CEDULA', how='left')
        final_df = pd.merge(merged_df, df_ac_arp, on='CEDULA', how='left')

        final_df['VALOR_POSITIVO'] = final_df['VALOR'].abs()
        resta_fs = final_df['ULTIMO_SALDO_FS'] - final_df['VALOR_POSITIVO']
        resta_arp = final_df['ULTIMO_SALDO_ARP'] - final_df['VALOR_POSITIVO']
        final_df['RESTA_SALDO'] = resta_fs.fillna(resta_arp)

        condiciones = [
            (pd.notna(final_df['FACTURA_FS'])) & (pd.notna(final_df['FACTURA_ARP'])),
            (final_df['RESTA_SALDO'] <= 0),
            (pd.notna(final_df['FACTURA_FS'])) & (pd.isna(final_df['FACTURA_ARP'])),
            (pd.isna(final_df['FACTURA_FS'])) & (pd.notna(final_df['FACTURA_ARP']))
        ]
        opciones = ['REVISAR TIENE 2 CARTERAS', 'PAGO TOTAL', 'CARTERA EN FINANSUEÑOS', 'CARTERA EN ARPESOD']
        final_df['OBSERVACIONES'] = np.select(condiciones, opciones, default='REVISAR SI ES CODEUDOR')
        
        multi_cartera_mask = final_df['OBSERVACIONES'] == 'REVISAR TIENE 2 CARTERAS'
        final_df.loc[multi_cartera_mask, ['FACTURA_FS', 'FACTURA_ARP']] = 'MAS DE UNA CARTERA'
        
        return final_df   
    
    def _prepare_output_sheets(self, final_df: pd.DataFrame):
        """Prepara y reindexa los DataFrames para cada hoja del reporte final."""
        df_finansuenos = final_df[pd.notna(final_df['FACTURA_FS'])].copy()
        df_arpesod = final_df[pd.notna(final_df['FACTURA_ARP'])].copy()
        df_sin_cartera = final_df[(pd.isna(final_df['FACTURA_FS'])) & (pd.isna(final_df['FACTURA_ARP']))].copy()
        
        # Corrección: Añadimos el reindex para la hoja 'SIN CARTERA' que faltaba
        base_cols = list(self.config.rename_columns['ONLINE'].values())
        cols_sin_cartera = base_cols + ['CUENTAS_FS', 'CUENTAS_ARP', 'OBSERVACIONES']
        
        return {
            "FINANSUEÑOS": df_finansuenos.reindex(columns=self.config.column_order_fs),
            "ARPESOD": df_arpesod.reindex(columns=self.config.column_order_arp),
            "SIN CARTERA": df_sin_cartera.reindex(columns=cols_sin_cartera)
        }
        
    def save_formatted_excel(self, output_path: str, sheets_data: Dict[str, pd.DataFrame]):
        """
        Guarda las hojas de datos YA PREPARADAS en un archivo Excel y les aplica formato.
        No vuelve a procesar nada.
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_data.items():
                    if df.empty:
                        continue # No guardar hojas vacías
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    self._format_sheet(worksheet, df) # Llama a un helper de formato
        except Exception as e:
            raise ValueError(f"Error al guardar el archivo Excel formateado:\n {str(e)}")     

    def _format_sheet(self, worksheet, df):
     
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name='Calibri')
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border_side = Side(style='thin', color='000000')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # 2. Formatear la fila del encabezado
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
            
        # 3. Calcular duplicados directamente desde el DataFrame
        # El método ahora es autocontenido y no necesita un 'duplicates_config'
        dup_cedulas = set(df[df.duplicated(subset=['CEDULA'], keep=False)]['CEDULA']) if 'CEDULA' in df.columns else set()
        dup_facturas_fs = set(df[df.duplicated(subset=['FACTURA_FS'], keep=False)]['FACTURA_FS']) if 'FACTURA_FS' in df.columns else set()
        dup_facturas_arp = set(df[df.duplicated(subset=['FACTURA_ARP'], keep=False)]['FACTURA_ARP']) if 'FACTURA_ARP' in df.columns else set()

        # 4. Iterar sobre las filas de datos para aplicar formato condicional
        col_indices = {cell.value: i for i, cell in enumerate(worksheet[1], 1)}

        for row_idx, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
            # Obtener valores clave de la fila actual
            obs_val = row_values[col_indices.get('OBSERVACIONES', 0) - 1] if 'OBSERVACIONES' in col_indices else None
            
            # Aplicar formato de fondo a toda la fila
            if obs_val == 'REVISAR TIENE 2 CARTERAS':
                for cell in worksheet[row_idx]: cell.fill = light_red_fill
            elif obs_val == 'PAGO TOTAL':
                for cell in worksheet[row_idx]: cell.fill = yellow_fill

            # Aplicar formato de fondo a celdas específicas (duplicados)
            cedula_val = row_values[col_indices.get('CEDULA', 0) - 1] if 'CEDULA' in col_indices else None
            if cedula_val in dup_cedulas:
                worksheet.cell(row=row_idx, column=col_indices['CEDULA']).fill = light_red_fill
            
            factura_fs_val = row_values[col_indices.get('FACTURA_FS', 0) - 1] if 'FACTURA_FS' in col_indices else None
            if factura_fs_val and factura_fs_val in dup_facturas_fs:
                worksheet.cell(row=row_idx, column=col_indices['FACTURA_FS']).fill = light_red_fill
            
            factura_arp_val = row_values[col_indices.get('FACTURA_ARP', 0) - 1] if 'FACTURA_ARP' in col_indices else None
            if factura_arp_val and factura_arp_val in dup_facturas_arp:
                worksheet.cell(row=row_idx, column=col_indices['FACTURA_ARP']).fill = light_red_fill

            # Aplicar bordes a todas las celdas de la fila
            for cell in worksheet[row_idx]:
                cell.border = cell_border

        # 5. Autoajustar el ancho de las columnas
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width