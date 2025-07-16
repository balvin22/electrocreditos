import pandas as pd
from typing import Dict
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class AnticiposReportWriter:
    """Guarda y formatea el reporte final de anticipos usando Openpyxl."""

    def save_report(self, output_path: str, sheets_data: Dict[str, pd.DataFrame]):
        """Guarda las hojas de datos en un archivo Excel y les aplica formato."""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_data.items():
                    if df.empty:
                        continue
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    self._format_sheet(worksheet, df)
        except Exception as e:
            raise ValueError(f"Error al guardar el archivo Excel formateado:\n {e}")

    def _format_sheet(self, worksheet, df: pd.DataFrame):
        """Aplica estilos de celda, condicionales y de encabezado a una hoja."""
        # 1. Definir estilos
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name='Calibri')
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border_side = Side(style='thin', color='000000')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # 2. Formatear encabezado
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
            
        # 3. Calcular duplicados
        dup_cedulas = set(df[df.duplicated(subset=['CEDULA'], keep=False)]['CEDULA']) if 'CEDULA' in df.columns else set()
        dup_facturas_fs = set(df[df.duplicated(subset=['FACTURA_FS'], keep=False)]['FACTURA_FS']) if 'FACTURA_FS' in df.columns else set()
        dup_facturas_arp = set(df[df.duplicated(subset=['FACTURA_ARP'], keep=False)]['FACTURA_ARP']) if 'FACTURA_ARP' in df.columns else set()

        # 4. Formatear filas de datos
        col_indices = {cell.value: i for i, cell in enumerate(worksheet[1], 1)}
        for row_idx, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
            obs_val = row_values[col_indices.get('OBSERVACIONES', 0) - 1] if 'OBSERVACIONES' in col_indices else None
            
            # Formato de fila completa
            if obs_val == 'REVISAR TIENE 2 CARTERAS':
                for cell in worksheet[row_idx]: cell.fill = light_red_fill
            elif obs_val == 'PAGO TOTAL':
                for cell in worksheet[row_idx]: cell.fill = yellow_fill

            # Formato de celdas específicas (duplicados)
            cedula_val = row_values[col_indices.get('CEDULA', 0) - 1] if 'CEDULA' in col_indices else None
            if cedula_val in dup_cedulas:
                worksheet.cell(row=row_idx, column=col_indices['CEDULA']).fill = light_red_fill
            
            # ... (Lógica para FACTURA_FS y FACTURA_ARP) ...

            # Aplicar bordes
            for cell in worksheet[row_idx]:
                cell.border = cell_border

        # 5. Autoajustar columnas
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width